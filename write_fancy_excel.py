from datetime import datetime
import psycopg2
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# DB Connection
conn = psycopg2.connect('host=localhost port=5437 user=postgres password=password')
curs = conn.cursor()

# Cubing query (postgres feature)
curs.execute("""
with mycube as (
select (date_part('year', payment_date) * 100 + date_part('month', payment_date))::int year_month,
       payment_date::date as payment_date,
       a.address, f.rating, sum(p.amount) as tot_payments, count(*) as num_payments,
       grouping((date_part('year', payment_date) * 100 + date_part('month', payment_date))::int, payment_date::date, a.address, f.rating) as mygrp
  from rental r
       join inventory i on i.inventory_id = r.inventory_id
       join film f on f.film_id = i.film_id
       join staff st on st.staff_id = r.staff_id
       join store s on s.store_id = st.store_id
       join address a on a.address_id = s.address_id
       join payment p on p.rental_id = r.rental_id
 group by cube (year_month, payment_date::date, a.address, f.rating)
)
select year_month, payment_date, address, rating, tot_payments, num_payments
  from mycube
 where mygrp in (15, 12, 0)
 order by mygrp desc, year_month nulls first, payment_date nulls first, address nulls first, rating nulls first
 ;
""")

# Create an excel workbook using OpenPyxl
wb = Workbook()
ws = wb.active

# Fill top summary of grand totals
r = curs.fetchone()
ws.append(['Report Name', 'Payments by Date, Store, Rating'])
ws.append(['Report Date', datetime.now()])
ws.append(['Total Payment Amount', r[4]])
ws.append(['Total Payment Count', r[5]])
ws.append([])


tables = list()
# Fill short summary of totals by store and rating
ws.append(['Store', 'Rating', 'Amount', 'Number'])
tables.append((ws.max_column, ws.max_row))
r = curs.fetchone()
while r:
    if r[0]: break  # Move to next section when the yyyymm values start appearing
    ws.append(r[2:])
    r = curs.fetchone()
tables.append((ws.max_column, ws.max_row))
ws.append([])

# Write headers for detail data and fill the rows
ws.append([d[0] for d in curs.description])
tables.append((ws.max_column, ws.max_row))
while r:
    ws.append(r)
    r = curs.fetchone()
tables.append((ws.max_column, ws.max_row))

# Make tables
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False,
        showLastColumn=False, showColumnStripes=False)
tabnum = 1
i = iter(tables)
for t in i:
    rng = f"A{t[1]}"
    t = next(i)
    rng = f"{rng}:{get_column_letter(t[0])}{t[1]}"
    tab = Table(displayName=f"Table{tabnum}", ref=rng)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    tabnum += 1

# Save the workbook
wb.save('SampleTables.xlsx')